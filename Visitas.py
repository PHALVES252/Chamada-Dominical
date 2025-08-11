import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime
import os

def validar_data(data_str):
    try:
        datetime.strptime(data_str, "%d/%m/%Y")
        return True
    except ValueError:
        return False

def salvar_visita():
    nome = entrada_nome.get().strip()
    data = entrada_data.get().strip()

    if not nome:
        messagebox.showerror("Erro", "Digite o nome da pessoa.")
        return

    if not validar_data(data):
        messagebox.showerror("Erro", "Digite a data no formato DD/MM/AAAA.")
        return

    arquivo = "registro_visitas.xlsx"

    if os.path.exists(arquivo):
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
        except InvalidFileException:
            messagebox.showerror("Erro", "O arquivo registro_visitas.xlsx está corrompido ou inválido.")
            return
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Visitas"
        ws.append(["Nome", "Data", "Horário"])

    # Adiciona o registro
    horario = datetime.now().strftime("%H:%M:%S")
    ws.append([nome, data, horario])

    try:
        wb.save(arquivo)
    except PermissionError:
        messagebox.showerror("Erro", "Feche o arquivo Excel antes de salvar.")
        return

    entrada_nome.delete(0, tk.END)
    entrada_data.delete(0, tk.END)
    messagebox.showinfo("Sucesso", f"Visita de {nome} registrada com sucesso!")

# Interface gráfica
root = tk.Tk()
root.title("Registro de Visitas")

tk.Label(root, text="Nome da pessoa:", font=("Arial", 12)).pack(pady=5)
entrada_nome = tk.Entry(root, width=40)
entrada_nome.pack()

tk.Label(root, text="Data da visita (DD/MM/AAAA):", font=("Arial", 12)).pack(pady=5)
entrada_data = tk.Entry(root, width=15)
entrada_data.pack()

tk.Button(root, text="Salvar Registro", command=salvar_visita, bg="green", fg="white", font=("Arial", 12)).pack(pady=15)

root.mainloop()
