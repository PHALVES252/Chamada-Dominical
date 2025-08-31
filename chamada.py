import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime
import os

alunos = [
    "OsValdina Francisca", "Paulo Henrique", "João Vitor", "Elza Alves",
    "Antonio Patricio", "Gesmindo Boostel", "Kalahan Boostel", "Geciel Polegario",
    "Diana", "Vanuza Nascimento", "Welington Nascimento",
    "Jorge", "Gosmira", "Almir Rodrigues","Herminio","Maria do Carmo","Laide"
]

# normaliza nomes (remove espaços no início/fim)
alunos = [a.strip() for a in alunos]

checkboxes = {}
widgets_check = {}  # guarda os widgets para conseguir destruir depois

def validar_data(data_str):
    try:
        return datetime.strptime(data_str, "%d/%m/%Y")
    except ValueError:
        return None

def salvar_chamada():
    data_str = entrada_data.get().strip()
    data_dt = validar_data(data_str)
    if not data_dt:
        messagebox.showerror("Erro", "Digite a data no formato DD/MM/AAAA.")
        return

    data_fmt = data_dt.strftime("%d/%m/%Y")
    arquivo = "chamada.xlsx"

    # Abre ou cria o arquivo
    if os.path.exists(arquivo):
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
        except InvalidFileException:
            messagebox.showerror("Erro", "O arquivo chamada.xlsx está corrompido ou inválido.")
            return
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Chamada"
        ws.append(["Nome", "Presença", "Data"])

    # coleta registros existentes
    registros_existentes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        nome_row = str(row[0]).strip()
        data_registro = row[2] if len(row) > 2 else None
        if isinstance(data_registro, datetime):
            data_row = data_registro.strftime("%d/%m/%Y")
        else:
            data_row = str(data_registro).strip() if data_registro is not None else ""
        registros_existentes.add((nome_row, data_row))

    novos = 0
    for nome, var in checkboxes.items():
        status = "Presente" if var.get() else "Ausente"
        if (nome, data_fmt) not in registros_existentes:
            ws.append([nome, status, data_fmt])
            novos += 1

    wb.save(arquivo)

    for var in checkboxes.values():
        var.set(False)

    messagebox.showinfo("Sucesso", f"Chamada registrada com sucesso! ({novos} novos registros adicionados)")

def excluir_aluno():
    nome_excluir = entrada_excluir.get().strip()
    if not nome_excluir:
        messagebox.showwarning("Aviso", "Digite um nome para excluir.")
        return

    if nome_excluir not in alunos:
        messagebox.showerror("Erro", f"O aluno '{nome_excluir}' não foi encontrado.")
        return

    # remove da lista
    alunos.remove(nome_excluir)

    # destrói o widget
    if nome_excluir in widgets_check:
        widgets_check[nome_excluir].destroy()
        del widgets_check[nome_excluir]
        del checkboxes[nome_excluir]

    messagebox.showinfo("Sucesso", f"O aluno '{nome_excluir}' foi removido.")

# Interface
root = tk.Tk()
root.title("Chamada Escolar")

tk.Label(root, text="Marque os alunos presentes:", font=("Arial", 14)).pack(pady=10)

frame = tk.Frame(root)
frame.pack()

for nome in alunos:
    var = tk.BooleanVar()
    chk = tk.Checkbutton(frame, text=nome, variable=var)
    chk.pack(anchor='w')
    checkboxes[nome] = var
    widgets_check[nome] = chk

frame_data = tk.Frame(root)
frame_data.pack(pady=10)

tk.Label(frame_data, text="Data da chamada (DD/MM/AAAA):").pack(side=tk.LEFT)
entrada_data = tk.Entry(frame_data)
entrada_data.pack(side=tk.LEFT)

tk.Button(root, text="Salvar Chamada", command=salvar_chamada, bg="green", fg="white").pack(pady=10)

# Área de exclusão
frame_excluir = tk.Frame(root)
frame_excluir.pack(pady=10)

tk.Label(frame_excluir, text="Excluir aluno:").pack(side=tk.LEFT)
entrada_excluir = tk.Entry(frame_excluir)
entrada_excluir.pack(side=tk.LEFT)

tk.Button(frame_excluir, text="Excluir", command=excluir_aluno, bg="red", fg="white").pack(side=tk.LEFT, padx=5)

root.mainloop()
